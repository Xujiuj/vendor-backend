from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


MARK = "source(A)"


def norm(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, float):
        return format(Decimal(str(value)).normalize(), "f")
    return value


def text(value) -> str | None:
    value = norm(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    return str(value).strip()


def integer(value) -> str | None:
    value = text(value)
    if not value:
        return None
    return str(int(Decimal(value)))


def decimal(value) -> str | None:
    value = text(value)
    if not value:
        return None
    return format(Decimal(value).normalize(), "f")


def sql(value) -> str:
    value = text(value)
    if value is None:
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_int(value) -> str:
    value = integer(value)
    return "NULL" if value is None else value


def sql_decimal(value) -> str:
    value = decimal(value)
    return "NULL" if value is None else value


def worksheet_rows(ws):
    iterator = ws.iter_rows(values_only=True)
    headers = [text(header) for header in next(iterator)]
    for row in iterator:
        record = {
            headers[index]: norm(row[index]) if index < len(row) else None
            for index in range(len(headers))
            if headers[index]
        }
        if any(value is not None and text(value) for value in record.values()):
            yield record


def find_sheet(workbook, prefix: str):
    return next(ws for ws in workbook.worksheets if ws.title.startswith(prefix))


def insert_block(lines: list[str], table: str, columns: list[str], values: list[list[str]], updates: list[str]) -> None:
    if not values:
        return
    lines.append(f"INSERT INTO {table}")
    lines.append("(" + ", ".join(columns) + ")")
    lines.append("VALUES")
    for index, row in enumerate(values):
        suffix = "," if index < len(values) - 1 else ""
        lines.append("(" + ", ".join(row) + ")" + suffix)
    if updates:
        lines.append("ON DUPLICATE KEY UPDATE")
        for index, update in enumerate(updates):
            suffix = "," if index < len(updates) - 1 else ";"
            lines.append("    " + update + suffix)
    else:
        lines[-1] += ";"
    lines.append("")


def main() -> None:
    repo = Path(__file__).resolve().parents[1]
    workspace = repo.parent
    source = next(path for path in workspace.iterdir() if path.is_dir() and path.name.startswith("source"))
    output = repo / "script" / "sql" / "mysql" / "carbon_vendor_source_a_dimension_seed.sql"

    workbook_1 = load_workbook(next((source / "ALL").glob("1*.xlsx")), read_only=True, data_only=True)
    workbook_2 = load_workbook(next((source / "ALL").glob("2*.xlsx")), read_only=True, data_only=True)
    content_workbook = load_workbook(next((source / "Content(A)").glob("Content.xlsx")), read_only=True, data_only=True)

    lines: list[str] = [
        "-- Source(A) vendor-owned master-data seed.",
        "-- Boundary: vendor cloud only. Enterprise-local sheets 102/104/201/204/3/4/5/501-504 are intentionally excluded.",
        "",
        "SET FOREIGN_KEY_CHECKS = 0;",
        f"DELETE FROM cv_report_content_catalog WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_electricity_factor WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_electricity_factor_version WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_electricity_factor_scope WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_greenhouse_gas WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_base_year WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_emission_source_category WHERE remark = {sql(MARK)};",
        f"DELETE FROM cv_admin_division WHERE remark = {sql(MARK)};",
        "SET FOREIGN_KEY_CHECKS = 1;",
        "",
    ]

    admin_values: dict[str, list[str]] = {}
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_1, "101")), 1):
        code = text(row.get("行政区划代码"))
        name = text(row.get("行政区划"))
        if code and name:
            admin_values[code] = [sql(code), sql(name), "NULL", sql("province"), str(index * 10), sql("0"), sql(MARK)]
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_2, "202")), len(admin_values) + 1):
        code = text(row.get("行政区划代码"))
        name = text(row.get("行政区划"))
        if code and name and code not in admin_values:
            admin_values[code] = [sql(code), sql(name), "NULL", sql("factor-region"), str(index * 10), sql("0"), sql(MARK)]
    insert_block(
        lines,
        "cv_admin_division",
        ["division_code", "division_name", "parent_code", "level_type", "sort_order", "status", "remark"],
        list(admin_values.values()),
        [
            "division_name = VALUES(division_name)",
            "parent_code = VALUES(parent_code)",
            "level_type = VALUES(level_type)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_1, "103")), 1):
        category_code = text(row.get("SK_排放源分类"))
        business_key = text(row.get("BK_业务键")) or category_code
        name = text(row.get("GHG Protocol范围子类别")) or text(row.get("统一标准分类")) or business_key
        if not category_code:
            continue
        values.append(
            [
                sql(category_code),
                sql(business_key),
                sql(name),
                sql(text(row.get("Scope Category (GHG Protocol)"))),
                sql(text(row.get("GHG Protocol范围"))),
                sql(text(row.get("GHG Protocol范围子类别"))),
                sql(text(row.get("ISO 14064-1类别"))),
                sql(text(row.get("ISO 14064-1 Category"))),
                sql(text(row.get("ISO 14064-1类别描述"))),
                sql(text(row.get("GB/T 32150-2025范围分类"))),
                sql(text(row.get("GB/T 32150-2025子类别"))),
                "NULL",
                sql(text(row.get("生效日期"))),
                sql(text(row.get("失效日期"))),
                sql(text(row.get("是否当前")) or "Y"),
                sql(text(row.get("版本号")) or "1"),
                sql(text(row.get("统一标准分类"))),
                str(index * 10),
                sql("0"),
                sql(MARK),
            ]
        )
    insert_block(
        lines,
        "cv_emission_source_category",
        [
            "category_code",
            "business_key",
            "category_name",
            "category_name_en",
            "ghg_scope",
            "ghg_scope_category",
            "iso_category",
            "iso_category_en",
            "iso_category_description",
            "gb_scope_category",
            "gb_subcategory",
            "parent_code",
            "effective_date",
            "expire_date",
            "current_flag",
            "version_no",
            "standard_category",
            "sort_order",
            "status",
            "remark",
        ],
        values,
        [
            "business_key = VALUES(business_key)",
            "category_name = VALUES(category_name)",
            "category_name_en = VALUES(category_name_en)",
            "ghg_scope = VALUES(ghg_scope)",
            "ghg_scope_category = VALUES(ghg_scope_category)",
            "iso_category = VALUES(iso_category)",
            "iso_category_en = VALUES(iso_category_en)",
            "iso_category_description = VALUES(iso_category_description)",
            "gb_scope_category = VALUES(gb_scope_category)",
            "gb_subcategory = VALUES(gb_subcategory)",
            "parent_code = VALUES(parent_code)",
            "effective_date = VALUES(effective_date)",
            "expire_date = VALUES(expire_date)",
            "current_flag = VALUES(current_flag)",
            "version_no = VALUES(version_no)",
            "standard_category = VALUES(standard_category)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_1, "106")), 1):
        key = text(row.get("基准年Key"))
        year = integer(row.get("基准年"))
        if not key or not year:
            continue
        is_current = "1" if text(row.get("是否当前基准")) == "是" else "0"
        values.append([sql(key), year, is_current, sql(text(row.get("说明"))), sql(key), sql(str(year)), str(index * 10), sql("0"), sql(MARK)])
    insert_block(
        lines,
        "cv_base_year",
        ["base_year_key", "base_year", "is_current", "description", "factory_code", "factory_name", "sort_order", "status", "remark"],
        values,
        [
            "base_year = VALUES(base_year)",
            "is_current = VALUES(is_current)",
            "description = VALUES(description)",
            "factory_code = VALUES(factory_code)",
            "factory_name = VALUES(factory_name)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_2, "203")), 1):
        values.append([sql(text(row.get("对应因子版本"))), sql_int(row.get("年份")), str(index * 10), sql("0"), sql(MARK)])
    insert_block(
        lines,
        "cv_electricity_factor_version",
        ["factor_version", "effective_year", "sort_order", "status", "remark"],
        values,
        [
            "factor_version = VALUES(factor_version)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_2, "202")), 1):
        values.append(
            [
                sql(text(row.get("因子版本"))),
                sql(text(row.get("行政区划代码"))),
                sql(text(row.get("行政区划"))),
                sql(text(row.get("区域划分"))),
                sql_decimal(row.get("省级因子（kgCO2/kWh)")),
                sql_decimal(row.get("区域因子（kgCO2/kWh)")),
                sql_decimal(row.get("全国因子（kgCO2/kWh）")),
                sql_decimal(row.get("不包括市场化交易的非化石能源电量因子（kgCO2/kWh）")),
                sql_decimal(row.get("全国化石能源电力二氧化碳排放因子（kgCO2/kWh）")),
                str(index * 10),
                sql("0"),
                sql(MARK),
            ]
        )
    insert_block(
        lines,
        "cv_electricity_factor",
        [
            "factor_version",
            "division_code",
            "division_name",
            "region_name",
            "province_factor",
            "region_factor",
            "national_factor",
            "non_fossil_excluded_factor",
            "national_fossil_power_factor",
            "sort_order",
            "status",
            "remark",
        ],
        values,
        [
            "division_name = VALUES(division_name)",
            "region_name = VALUES(region_name)",
            "province_factor = VALUES(province_factor)",
            "region_factor = VALUES(region_factor)",
            "national_factor = VALUES(national_factor)",
            "non_fossil_excluded_factor = VALUES(non_fossil_excluded_factor)",
            "national_fossil_power_factor = VALUES(national_fossil_power_factor)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_2, "205")), 1):
        values.append([sql(text(row.get("因子口径Key"))), sql(text(row.get("因子口径"))), str(index * 10), sql("0"), sql(MARK)])
    insert_block(
        lines,
        "cv_electricity_factor_scope",
        ["scope_key", "scope_name", "sort_order", "status", "remark"],
        values,
        [
            "scope_name = VALUES(scope_name)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    for index, row in enumerate(worksheet_rows(find_sheet(workbook_2, "206")), 1):
        gas = text(row.get("气体"))
        if gas:
            values.append([sql(gas), sql(gas), sql(gas), "NULL", "NULL", sql(gas), sql_int(row.get("排序")) or str(index * 10), sql("0"), sql(MARK)])
    insert_block(
        lines,
        "cv_greenhouse_gas",
        ["gas_code", "gas_name", "gas_name_en", "gwp_value", "gwp_version", "chemical_formula", "sort_order", "status", "remark"],
        values,
        [
            "gas_name = VALUES(gas_name)",
            "gas_name_en = VALUES(gas_name_en)",
            "gwp_value = VALUES(gwp_value)",
            "gwp_version = VALUES(gwp_version)",
            "chemical_formula = VALUES(chemical_formula)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    values = []
    current_catalog_no = None
    current_catalog_name = None
    for index, row in enumerate(worksheet_rows(content_workbook["Sheet1"]), 1):
        if text(row.get("目录序号")):
            current_catalog_no = text(row.get("目录序号"))
        if text(row.get("目录")):
            current_catalog_name = text(row.get("目录"))
        subcatalog_no = text(row.get("子目录序号"))
        subcatalog_name = text(row.get("子目录"))
        if subcatalog_no and subcatalog_name:
            values.append(
                [
                    sql(current_catalog_no),
                    sql(current_catalog_name),
                    sql(subcatalog_no),
                    sql(subcatalog_name),
                    sql(text(row.get("页面图表"))),
                    str(index * 10),
                    sql("0"),
                    sql(MARK),
                ]
            )
    insert_block(
        lines,
        "cv_report_content_catalog",
        ["catalog_no", "catalog_name", "subcatalog_no", "subcatalog_name", "chart_list", "sort_order", "status", "remark"],
        values,
        [
            "catalog_no = VALUES(catalog_no)",
            "catalog_name = VALUES(catalog_name)",
            "subcatalog_name = VALUES(subcatalog_name)",
            "chart_list = VALUES(chart_list)",
            "sort_order = VALUES(sort_order)",
            "status = VALUES(status)",
            "remark = VALUES(remark)",
        ],
    )

    lines.extend(
        [
            "SELECT 'cv_admin_division' AS table_name, COUNT(*) AS source_a_rows FROM cv_admin_division WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_emission_source_category', COUNT(*) FROM cv_emission_source_category WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_base_year', COUNT(*) FROM cv_base_year WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_electricity_factor_version', COUNT(*) FROM cv_electricity_factor_version WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_electricity_factor', COUNT(*) FROM cv_electricity_factor WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_electricity_factor_scope', COUNT(*) FROM cv_electricity_factor_scope WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_greenhouse_gas', COUNT(*) FROM cv_greenhouse_gas WHERE remark = 'source(A)'",
            "UNION ALL SELECT 'cv_report_content_catalog', COUNT(*) FROM cv_report_content_catalog WHERE remark = 'source(A)';",
        ]
    )

    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
