from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
WORKSPACE = REPO.parent
SOURCE = WORKSPACE / "source（A）"
INIT_SQL = REPO / "script" / "sql" / "sqlserver" / "carbon_vendor_init.sql"
MARK = "source(A)"

START_MARKER = "    INSERT INTO dbo.cv_admin_division"
END_MARKER = "    COMMIT TRANSACTION;"
PACKAGE_START_MARKER = "    INSERT INTO dbo.sys_tenant_package"
PACKAGE_END_MARKER = "    INSERT INTO dbo.cv_signing_key"

GAS_EN = {
    "CO2": "Carbon dioxide",
    "CH4": "Methane",
    "N2O": "Nitrous oxide",
    "HFCs": "Hydrofluorocarbons",
    "PFCs": "Perfluorocarbons",
    "SF6": "Sulfur hexafluoride",
    "NF3": "Nitrogen trifluoride",
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float):
        return Decimal(str(value)).normalize()
    return value


def text(value: Any) -> str | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        return format(Decimal(str(value)).normalize(), "f")
    return str(value).strip()


def integer(value: Any) -> int | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    return int(Decimal(str(value)))


def decimal(value: Any) -> Decimal | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value)).normalize()


def rows(workbook_path: Path, sheet_prefix: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = next((ws for ws in workbook.worksheets if ws.title.startswith(sheet_prefix)), None)
    if sheet is None:
        raise RuntimeError(f"Missing sheet prefix {sheet_prefix!r} in {workbook_path}")
    header_row_no, header = header_row(sheet)
    result: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=header_row_no + 1, values_only=True):
        row = {
            str(header[index]): clean(raw[index]) if index < len(raw) else None
            for index in range(len(header))
            if header[index]
        }
        if any(value is not None for value in row.values()):
            result.append(row)
    return result


def all_content_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    result: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        try:
            header_row_no, header = header_row(sheet)
        except RuntimeError:
            continue
        for raw in sheet.iter_rows(min_row=header_row_no + 1, values_only=True):
            row = {
                str(header[index]): clean(raw[index]) if index < len(raw) else None
                for index in range(len(header))
                if header[index]
            }
            if any(value is not None for value in row.values()):
                result.append(row)
    return result


def header_row(sheet: Any) -> tuple[int, list[str | None]]:
    best_row_no = 0
    best_header: list[str | None] = []
    best_score = 0
    for row_no, raw in enumerate(sheet.iter_rows(max_row=30, values_only=True), start=1):
        header = [text(value) for value in raw]
        score = sum(1 for value in header if value)
        if score > best_score:
            best_row_no = row_no
            best_header = header
            best_score = score
    if best_score == 0:
        raise RuntimeError(f"Cannot find header row in sheet {sheet.title}")
    return best_row_no, best_header


def q(value: Any) -> str:
    value = clean(value)
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, Decimal)):
        return format(value, "f") if isinstance(value, Decimal) else str(value)
    if isinstance(value, float):
        return format(Decimal(str(value)).normalize(), "f")
    if isinstance(value, datetime):
        return "'" + value.strftime("%Y-%m-%dT%H:%M:%S") + "'"
    if isinstance(value, date):
        return "'" + value.strftime("%Y-%m-%d") + "'"
    escaped = str(value).replace("'", "''")
    return "N'" + escaped + "'"


def normalize_code(value: Any, fallback: str) -> str:
    raw = text(value) or fallback
    raw = raw.lower().replace("（", "-").replace("）", "")
    raw = re.sub(r"[^a-z0-9]+", "-", raw)
    return raw.strip("-") or fallback


def chart_json(value: Any) -> str:
    raw = text(value)
    if not raw:
        return "[]"
    items = [item.strip() for item in re.split(r"[、,，;；\n]+", raw) if item.strip()]
    return json.dumps(items or [raw], ensure_ascii=False)


def first_int(value: Any, fallback: int) -> int:
    raw = text(value)
    if not raw:
        return fallback
    match = re.search(r"\d+", raw)
    return int(match.group(0)) if match else fallback


def source_templates() -> list[dict[str, Any]]:
    files = [
        ("source-a-emission-source-identification", "Source(A) 排放源识别表", SOURCE / "ALL" / "1 排放源识别表.xlsx"),
        ("source-a-emission-factor", "Source(A) 排放因子表", SOURCE / "ALL" / "2 排放因子表.xlsx"),
        ("source-a-report-content", "Source(A) 报告内容模板", SOURCE / "Content(A)" / "Content.xlsx"),
        ("source-a-activity-10101", "Source(A) 排放活动数据表 10101", SOURCE / "活动数据表" / "3 排放活动数据表10101.xlsx"),
        ("source-a-activity-10102", "Source(A) 排放活动数据表 10102", SOURCE / "活动数据表" / "3 排放活动数据表10102.xlsx"),
        ("source-a-activity-10103", "Source(A) 排放活动数据表 10103", SOURCE / "活动数据表" / "3 排放活动数据表10103.xlsx"),
        ("source-a-activity-10104", "Source(A) 排放活动数据表 10104", SOURCE / "活动数据表" / "3 排放活动数据表10104.xlsx"),
        ("source-a-green-power", "Source(A) 绿证绿电表", SOURCE / "绿证绿电表" / "4 绿证绿电表.xlsx"),
        ("source-a-intensity-dimension", "Source(A) 碳排放强度维度表", SOURCE / "强度数据表" / "5碳排放强度维度表.xlsx"),
        ("source-a-denominator-fact", "Source(A) 分母事实表", SOURCE / "强度数据表" / "503分母事实表" / "503分母事实表.xlsx"),
    ]
    data = []
    for index, (code, name, path) in enumerate(files, start=1):
        if not path.exists():
            raise RuntimeError(f"Missing Source(A) template file: {path}")
        data.append({
            "id": index,
            "template_code": code,
            "template_name": name,
            "template_version": "source(A)",
            "file_name": path.name,
            "file_uri": "template://source(A)/" + str(path.relative_to(SOURCE)).replace("\\", "/"),
            "publish_status": "published",
            "published_by": "admin",
            "published_time": "@now",
            "create_time": "@now",
            "remark": MARK,
        })
    return data


def sql_value(value: Any) -> str:
    return value if isinstance(value, str) and value.startswith("@") else q(value)


def values_insert(table: str, columns: list[str], data: list[dict[str, Any]]) -> str:
    if not data:
        return ""
    lines = [
        f"    INSERT INTO dbo.{table}",
        "        (" + ", ".join(columns) + ")",
        "    VALUES",
    ]
    values = []
    for row in data:
        values.append("        (" + ", ".join(sql_value(row.get(column)) for column in columns) + ")")
    return "\n".join(lines + [",\n".join(values) + ";"])


def template_entitlements() -> str:
    entitlements = [
        {"templateCode": template["template_code"], "templateVersion": template["template_version"], "scope": "download"}
        for template in source_templates()
    ]
    return json.dumps(entitlements, ensure_ascii=False, separators=(",", ":"))


def package_block() -> str:
    entitlements = template_entitlements()
    return f"""    INSERT INTO dbo.sys_tenant_package
        (package_id, package_name, menu_ids, remark, menu_check_strictly, price_amount, price_currency, billing_cycle,
         online_purchase_enabled, license_auto_issue_enabled, license_key_id, license_validity_days, license_feature_codes,
         license_template_entitlements, status, del_flag, create_dept, create_by, create_time, update_by, update_time)
    VALUES
        (1001, N'基础套餐', @menuIds, N'适用于单组织基础填报和 Source(A) 模板下载。', 1, 1999.00, N'CNY', N'YEAR',
         1, 0, N'local-key-2026-06-08', NULL, N'capture,report-template-download',
         {q(entitlements)},
         N'0', N'0', @createDept, @createBy, @now, NULL, NULL),
        (1002, N'专业套餐', @menuIds, N'适用于多部门协同、因子同步和 Source(A) 报表模板同步。', 1, 0.00, N'CNY', N'YEAR',
         0, 0, N'local-key-2026-06-08', NULL, N'capture,factor-sync,report-template-sync,report-template-download',
         {q(entitlements)},
         N'0', N'0', @createDept, @createBy, @now, NULL, NULL),
        (1003, N'集团套餐', @menuIds, N'适用于集团客户、报表门禁和 Source(A) 全量模板能力。', 1, 0.00, N'CNY', N'YEAR',
         0, 0, N'local-key-2026-06-08', NULL, N'capture,factor-sync,report-template-sync,report-template-download,report-gate,premium-support',
         {q(entitlements)},
         N'0', N'0', @createDept, @createBy, @now, NULL, NULL);
"""


def factor_gwp_by_gas(ef201: Iterable[dict[str, Any]]) -> dict[str, Decimal]:
    result = {"CO2": Decimal("1")}
    source_columns = {
        "CH4": "GWP_CH4",
        "N2O": "GWP_N2O",
        "HFCs": "GWP_HFCs",
        "PFCs": "GWP_PFCs",
        "SF6": "GWP_SF6",
        "NF3": "GWP_NF3",
    }
    for row in ef201:
        for gas, column in source_columns.items():
            if gas not in result:
                value = decimal(row.get(column))
                if value is not None:
                    result[gas] = value
    return result


def build_seed_block() -> str:
    source_book = SOURCE / "ALL" / "1 排放源识别表.xlsx"
    factor_book = SOURCE / "ALL" / "2 排放因子表.xlsx"
    content_book = SOURCE / "Content(A)" / "Content.xlsx"

    admin = rows(source_book, "101")
    category = rows(source_book, "103")
    base_year = rows(source_book, "106")
    ef201 = rows(factor_book, "201")
    ef202 = rows(factor_book, "202")
    ef203 = rows(factor_book, "203")
    ef205 = rows(factor_book, "205")
    gas206 = rows(factor_book, "206")
    content = all_content_rows(content_book)

    blocks: list[str] = []
    blocks.append(values_insert("cv_admin_division", [
        "id", "division_code", "division_name", "parent_code", "level_type", "sort_order", "status",
        "create_dept", "create_by", "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index, "division_code": text(row["行政区划代码"]), "division_name": row["行政区划"],
        "parent_code": None, "level_type": "province", "sort_order": index, "status": "0",
        "create_dept": "@createDept", "create_by": "@createBy", "create_time": "@now",
        "update_by": None, "update_time": None, "remark": MARK,
    } for index, row in enumerate(admin, start=1)]))

    blocks.append(values_insert("cv_base_year", [
        "id", "base_year_key", "base_year", "is_current", "description", "sort_order", "status",
        "create_dept", "create_by", "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index,
        "base_year_key": text(row.get("基准年Key")) or text(row.get("基准年")),
        "base_year": integer(row.get("基准年")),
        "is_current": 1 if text(row.get("是否当前基准")) in {"是", "1", "Y", "y", "true", "True"} else 0,
        "description": row.get("说明"),
        "sort_order": index * 10,
        "status": "0",
        "create_dept": "@createDept",
        "create_by": "@createBy",
        "create_time": "@now",
        "update_by": None,
        "update_time": None,
        "remark": MARK,
    } for index, row in enumerate(base_year, start=1)]))

    category_rows = []
    for index, row in enumerate(category, start=1):
        business_key = text(row.get("BK_业务键")) or str(index)
        category_rows.append({
            "id": index,
            "category_code": text(row.get("SK_排放源分类")) or str(index),
            "business_key": business_key,
            "category_name": row.get("GHG Protocol范围子类别"),
            "category_name_en": row.get("Scope Category (GHG Protocol)"),
            "ghg_scope": row.get("GHG Protocol范围"),
            "ghg_scope_category_sort": first_int(row.get("GHG Protocol范围子类别排序"), index),
            "ghg_scope_category": row.get("GHG Protocol范围子类别"),
            "ghg_scope_en": row.get("Scope (GHG Protocol)"),
            "ghg_scope_category_en": row.get("Scope Category (GHG Protocol)"),
            "iso_category": row.get("ISO 14064-1类别"),
            "iso_category_en": row.get("ISO 14064-1 Category"),
            "iso_category_description": row.get("ISO 14064-1类别描述"),
            "iso_category_description_en": row.get("ISO 14064-1 Category Description (EN)"),
            "iso_custom_subcategory": row.get("ISO 14064-1子类别（自定义）"),
            "gb_scope_category": row.get("GB/T 32150-2025范围分类"),
            "gb_subcategory": row.get("GB/T 32150-2025子类别"),
            "parent_code": None,
            "effective_date": row.get("生效日期"),
            "expire_date": row.get("失效日期"),
            "current_flag": "Y" if text(row.get("是否当前")) in {"是", "1", "Y", "y", "true", "True"} else "N",
            "version_no": text(row.get("版本号")),
            "standard_category": row.get("统一标准分类"),
            "category_current_key": business_key,
            "sort_order": index,
            "status": "0",
            "create_dept": "@createDept",
            "create_by": "@createBy",
            "create_time": "@now",
            "update_by": None,
            "update_time": None,
            "remark": MARK,
        })
    blocks.append(values_insert("cv_emission_source_category", [
        "id", "category_code", "business_key", "category_name", "category_name_en", "ghg_scope",
        "ghg_scope_category_sort", "ghg_scope_category", "ghg_scope_en", "ghg_scope_category_en",
        "iso_category", "iso_category_en", "iso_category_description", "iso_category_description_en",
        "iso_custom_subcategory", "gb_scope_category", "gb_subcategory", "parent_code", "effective_date",
        "expire_date", "current_flag", "version_no",
        "standard_category", "category_current_key", "sort_order", "status", "create_dept", "create_by",
        "create_time", "update_by", "update_time", "remark"
    ], category_rows))

    blocks.append(values_insert("cv_electricity_factor_scope", [
        "id", "scope_key", "scope_name", "sort_order", "status", "create_dept", "create_by",
        "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index, "scope_key": text(row.get("因子口径Key")), "scope_name": row.get("因子口径"),
        "sort_order": index, "status": "0", "create_dept": "@createDept", "create_by": "@createBy",
        "create_time": "@now", "update_by": None, "update_time": None, "remark": MARK,
    } for index, row in enumerate(ef205, start=1)]))

    blocks.append(values_insert("cv_electricity_factor_version", [
        "id", "factor_version", "effective_year", "sort_order", "status", "create_dept", "create_by",
        "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index, "factor_version": text(row.get("对应因子版本")), "effective_year": integer(row.get("年份")), "sort_order": index,
        "status": "0", "create_dept": "@createDept", "create_by": "@createBy", "create_time": "@now",
        "update_by": None, "update_time": None, "remark": MARK,
    } for index, row in enumerate(ef203, start=1) if text(row.get("对应因子版本")) and integer(row.get("年份")) is not None]))

    blocks.append(values_insert("cv_electricity_factor", [
        "id", "version_province_code", "factor_version", "division_code", "division_name", "region_name", "province_factor", "region_factor",
        "national_factor", "non_fossil_excluded_factor", "national_fossil_power_factor", "sort_order", "status",
        "create_dept", "create_by", "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index, "version_province_code": text(row.get("PK_因子版本省份代码")), "factor_version": row.get("因子版本"), "division_code": text(row.get("行政区划代码")),
        "division_name": row.get("行政区划"), "region_name": row.get("区域划分"),
        "province_factor": row.get("省级因子（kgCO2/kWh)"), "region_factor": row.get("区域因子（kgCO2/kWh)"),
        "national_factor": row.get("全国因子（kgCO2/kWh）"),
        "non_fossil_excluded_factor": row.get("不包括市场化交易的非化石能源电量因子（kgCO2/kWh）"),
        "national_fossil_power_factor": row.get("全国化石能源电力二氧化碳排放因子（kgCO2/kWh）"),
        "sort_order": index, "status": "0", "create_dept": "@createDept", "create_by": "@createBy",
        "create_time": "@now", "update_by": None, "update_time": None, "remark": MARK,
    } for index, row in enumerate(ef202, start=1)]))

    gwp = factor_gwp_by_gas(ef201)
    blocks.append(values_insert("cv_greenhouse_gas", [
        "id", "gas_code", "gas_name", "gas_name_en", "gwp_value", "gwp_version", "chemical_formula",
        "sort_order", "status", "create_dept", "create_by", "create_time", "update_by", "update_time", "remark"
    ], [{
        "id": index, "gas_code": row.get("GasKey"), "gas_name": row.get("气体"),
        "gas_name_en": GAS_EN.get(text(row.get("GasKey")) or ""), "gwp_value": gwp.get(text(row.get("GasKey")) or ""),
        "gwp_version": MARK, "chemical_formula": row.get("GasKey"), "sort_order": integer(row.get("排序")) or index,
        "status": "0", "create_dept": "@createDept", "create_by": "@createBy", "create_time": "@now",
        "update_by": None, "update_time": None, "remark": MARK,
    } for index, row in enumerate(gas206, start=1)]))

    blocks.append(values_insert("cv_factor_version", [
        "id", "version_code", "version_name", "publish_status", "frozen_flag", "published_by",
        "published_time", "create_time", "remark"
    ], [{
        "id": 1, "version_code": "source-a", "version_name": "Source(A) 排放因子库",
        "publish_status": "published", "frozen_flag": 1, "published_by": "admin",
        "published_time": "@now", "create_time": "@now", "remark": MARK,
    }]))

    factor_records = factor_record_rows(ef201, ef202)
    factor_columns = [
        "id", "version_id", "factor_table_code", "factor_code", "factor_name", "factor_category", "factor_value",
        "factor_unit", "factor_key", "emission_source_name", "emission_source_name_en", "fuel_material_category",
        "source_unit", "co2", "ch4", "n2o", "hfcs", "pfcs", "sf6", "nf3", "applicable_scope", "factor_source",
        "gwp_ch4", "gwp_n2o", "gwp_hfcs", "gwp_pfcs", "gwp_sf6", "gwp_nf3", "factor_gwp",
        "version_province_code", "factor_version", "division_code", "division_name", "region_name",
        "province_factor", "region_factor", "national_factor", "non_fossil_excluded_factor",
        "national_fossil_power_factor", "row_no", "fuel_level1", "fuel_level2", "fuel_level3", "fuel_level4",
        "lower_heat_value", "lower_heat_value_cv", "co2_factor", "co2_factor_cv", "gwp_value",
        "converted_factor", "source_ref", "custom_fields", "enabled_flag", "create_time", "update_time", "remark"
    ]
    blocks.append(values_insert("cv_factor_record", factor_columns, factor_records))

    blocks.append(values_insert("cv_factor_customer_scope", [
        "id", "version_id", "package_id", "package_name", "scope_status", "create_time"
    ], [
        {"id": 1, "version_id": 1, "package_id": 1001, "package_name": "基础套餐", "scope_status": "enabled", "create_time": "@now"},
        {"id": 2, "version_id": 1, "package_id": 1002, "package_name": "专业套餐", "scope_status": "enabled", "create_time": "@now"},
        {"id": 3, "version_id": 1, "package_id": 1003, "package_name": "集团套餐", "scope_status": "enabled", "create_time": "@now"},
    ]))

    templates = source_templates()
    blocks.append(values_insert("cv_report_template", [
        "id", "template_code", "template_name", "template_version", "file_name", "file_uri", "publish_status",
        "published_by", "published_time", "create_time", "remark"
    ], templates))
    template_scopes = []
    scope_id = 1
    for template in templates:
        for package_id, package_name in [(1001, "基础套餐"), (1002, "专业套餐"), (1003, "集团套餐")]:
            template_scopes.append({
                "id": scope_id, "template_id": template["id"], "package_id": package_id,
                "package_name": package_name, "scope_status": "enabled", "create_time": "@now",
            })
            scope_id += 1
    blocks.append(values_insert("cv_report_template_scope", [
        "id", "template_id", "package_id", "package_name", "scope_status", "create_time"
    ], template_scopes))

    content_rows = []
    catalog_rows = []
    current_directory_no: int | None = None
    current_directory_catalog: str | None = None
    current_directory_name: str | None = None
    for index, row in enumerate(content, start=1):
        raw_directory_no = text(row.get("目录序号"))
        raw_directory_name = text(row.get("目录"))
        if raw_directory_no:
            current_directory_no = first_int(raw_directory_no, index)
            current_directory_catalog = raw_directory_no
        if raw_directory_name:
            current_directory_name = raw_directory_name
        directory_no = current_directory_no or index
        directory_catalog = current_directory_catalog or str(directory_no)
        directory_name = current_directory_name or "未分组"
        subdirectory_no = first_int(row.get("子目录序号"), index)
        charts = chart_json(row.get("页面图表"))
        content_rows.append({
            "id": index, "directory_no": directory_no, "directory_name": directory_name,
            "subdirectory_no": subdirectory_no, "subdirectory_name": row.get("子目录"),
            "chart_names": charts, "display_order": index, "status": "0",
            "create_time": "@now", "update_time": None, "remark": MARK,
        })
        catalog_rows.append({
            "id": index, "catalog_no": directory_catalog, "catalog_name": directory_name,
            "subcatalog_no": text(row.get("子目录序号")), "subcatalog_name": row.get("子目录"),
            "chart_list": charts, "sort_order": index, "status": "0",
            "create_dept": "@createDept", "create_by": "@createBy", "create_time": "@now",
            "update_by": None, "update_time": None, "remark": MARK,
        })
    blocks.append(values_insert("cv_report_content", [
        "directory_no", "directory_name", "subdirectory_no", "subdirectory_name", "chart_names",
        "display_order", "status", "create_time", "update_time", "remark"
    ], content_rows))
    blocks.append(values_insert("cv_report_content_catalog", [
        "id", "catalog_no", "catalog_name", "subcatalog_no", "subcatalog_name", "chart_list", "sort_order",
        "status", "create_dept", "create_by", "create_time", "update_by", "update_time", "remark"
    ], catalog_rows))

    return "\n\n".join(blocks)


def factor_record_rows(ef201: list[dict[str, Any]], ef202: list[dict[str, Any]]) -> list[dict[str, Any]]:
    factor_records = []
    next_id = 1
    for row in ef201:
        factor_records.append({
            "id": next_id, "version_id": 1, "factor_table_code": "201ef",
            "factor_code": text(row.get("SK_排放因子")), "factor_name": row.get("排放源"),
            "factor_category": row.get("燃料/物质类别"), "factor_value": row.get("因子GWP"),
            "factor_unit": row.get("因子单位"), "factor_key": text(row.get("SK_排放因子")),
            "emission_source_name": row.get("排放源"), "emission_source_name_en": row.get("排放源_EN"),
            "fuel_material_category": row.get("燃料/物质类别"), "source_unit": row.get("排放源单位"),
            "co2": row.get("CO2"), "ch4": row.get("CH4"), "n2o": row.get("N2O"),
            "hfcs": row.get("HFCs"), "pfcs": row.get("PFCs"), "sf6": row.get("SF6"), "nf3": row.get("NF3"),
            "applicable_scope": row.get("适用范围"), "factor_source": row.get("因子来源"),
            "gwp_ch4": row.get("GWP_CH4"), "gwp_n2o": row.get("GWP_N2O"), "gwp_hfcs": row.get("GWP_HFCs"),
            "gwp_pfcs": row.get("GWP_PFCs"), "gwp_sf6": row.get("GWP_SF6"), "gwp_nf3": row.get("GWP_NF3"),
            "factor_gwp": row.get("因子GWP"), "version_province_code": None, "factor_version": None,
            "division_code": None, "division_name": None, "region_name": None, "province_factor": None,
            "region_factor": None, "national_factor": None, "non_fossil_excluded_factor": None,
            "national_fossil_power_factor": None, "row_no": next_id, "fuel_level1": None, "fuel_level2": None,
            "fuel_level3": None, "fuel_level4": None, "lower_heat_value": None, "lower_heat_value_cv": None,
            "co2_factor": None, "co2_factor_cv": None, "gwp_value": row.get("因子GWP"),
            "converted_factor": row.get("因子GWP"), "source_ref": "201EF排放因子维度表",
            "custom_fields": None, "enabled_flag": 1, "create_time": "@now", "update_time": None, "remark": MARK,
        })
        next_id += 1
    for row in ef202:
        value = row.get("省级因子（kgCO2/kWh)") or row.get("区域因子（kgCO2/kWh)") or row.get("全国因子（kgCO2/kWh）")
        version_province_code = text(row.get("PK_因子版本省份代码")) or f"{text(row.get('因子版本'))}{text(row.get('行政区划代码'))}"
        factor_records.append({
            "id": next_id, "version_id": 1, "factor_table_code": "202ef",
            "factor_code": version_province_code,
            "factor_name": row.get("行政区划"), "factor_category": "ef-electricity-factor",
            "factor_value": value, "factor_unit": "kgCO2/kWh",
            "factor_key": version_province_code,
            "emission_source_name": "电力", "emission_source_name_en": "Electricity",
            "fuel_material_category": "电力", "source_unit": "kWh", "co2": value, "ch4": 0, "n2o": 0,
            "hfcs": None, "pfcs": None, "sf6": None, "nf3": None, "applicable_scope": "电力因子",
            "factor_source": MARK, "gwp_ch4": None, "gwp_n2o": None, "gwp_hfcs": None, "gwp_pfcs": None,
            "gwp_sf6": None, "gwp_nf3": None, "factor_gwp": value,
            "version_province_code": version_province_code,
            "factor_version": text(row.get("因子版本")), "division_code": text(row.get("行政区划代码")),
            "division_name": row.get("行政区划"), "region_name": row.get("区域划分"),
            "province_factor": row.get("省级因子（kgCO2/kWh)"), "region_factor": row.get("区域因子（kgCO2/kWh)"),
            "national_factor": row.get("全国因子（kgCO2/kWh）"),
            "non_fossil_excluded_factor": row.get("不包括市场化交易的非化石能源电量因子（kgCO2/kWh）"),
            "national_fossil_power_factor": row.get("全国化石能源电力二氧化碳排放因子（kgCO2/kWh）"),
            "row_no": next_id, "fuel_level1": None, "fuel_level2": None, "fuel_level3": None, "fuel_level4": None,
            "lower_heat_value": None, "lower_heat_value_cv": None, "co2_factor": None, "co2_factor_cv": None,
            "gwp_value": value, "converted_factor": value, "source_ref": "202EF电力因子维度表",
            "custom_fields": None, "enabled_flag": 1, "create_time": "@now", "update_time": None, "remark": MARK,
        })
        next_id += 1
    return factor_records


def replace_seed_block(seed_block: str) -> None:
    sql = INIT_SQL.read_text(encoding="utf-8")
    package_start = sql.index(PACKAGE_START_MARKER)
    package_end = sql.index(PACKAGE_END_MARKER, package_start)
    sql = sql[:package_start] + package_block().rstrip() + "\n\n" + sql[package_end:]
    start = sql.index(START_MARKER)
    end = sql.index(END_MARKER, start)
    INIT_SQL.write_text(sql[:start] + seed_block.rstrip() + "\n\n" + sql[end:], encoding="utf-8", newline="\n")


def main() -> None:
    replace_seed_block(build_seed_block())
    print(f"Updated {INIT_SQL}")


if __name__ == "__main__":
    main()
